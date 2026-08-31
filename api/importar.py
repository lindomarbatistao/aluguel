from rest_framework.views import APIView
from openpyxl import load_workbook
from rest_framework.response import Response
from .models import *
from rest_framework import status

class ImportarUsuarios(APIView):
    
    def post(self, request):
        arquivo = request.FILES.get("file")
        
        planilha = load_workbook(arquivo)
        
        pagina = planilha["Usuarios"]
        
        # print(pagina)
        
        # for linha in pagina.iter_rows(values_only=True):
        #     print(linha)
        
        criados = 0 
        ignorados = 0
        
        for coluna in pagina.iter_rows(min_row=2,values_only=True):
            username=coluna[0]
            first_name=coluna[1]
            last_name=coluna[2]
            email=coluna[3]
            phone=coluna[4]
            type=coluna[5]
            is_active=coluna[6]
            senha_teste=coluna[7]
            
            if Usuario.objects.filter(username=username).exists():
                ignorados += 1
                continue
            
            usuario = Usuario(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                celular=phone,
                tipo=type,
                is_active=is_active,
                password=senha_teste,
            )
            
            usuario.set_password(senha_teste)

            usuario.save()
            
            criados += 1
            
        return Response(
            {
                "mensagem":"Planilha lida com sucesso!",
                "criados":criados,
                "ignorados": ignorados
            }, status=status.HTTP_201_CREATED
            )
        
class ImportarImoveis(APIView):
    def post(self, request):
        arquivo = request.FILES.get("file")
        planilha = load_workbook(arquivo)
        pagina = planilha["Imoveis"]
        criados = 0
        ignorados = 0
        
        for coluna in pagina.iter_rows(
            min_row=2,
            values_only=True
        ):
            title = coluna[0]
            type = coluna[1]
            rent_value = coluna[2]
            status = coluna[3]
            street_address = coluna[4]
            cep = coluna[5]
            complement = coluna[6]
            neighborhood = coluna[7]
            city = coluna[8]
            uf = coluna[9]

            if Imovel.objects.filter(titulo=title).exists():
                ignorados += 1
                continue
            
            imovel = Imovel(
                titulo = title,
                tipo = type,
                valor_aluguel = rent_value,
                status = status,
                logradouro = street_address,
                cep = cep,
                complemento = complement,
                bairro = neighborhood,
                cidade = city,
                uf = uf
            )
            
            imovel.save()
            criados += 1
            
        return Response(
            {
                "mensagem":"Planilha lida com sucesso!",
                "criados":criados,
                "ignorados": ignorados
            }
            )
            